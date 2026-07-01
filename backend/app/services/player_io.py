"""Player import (CSV/Excel) and export (CSV/Excel)."""

from __future__ import annotations

import csv
import io
import uuid
from datetime import date, datetime

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.enums import Gender, PlayerStatus
from app.models.org import Club, StateAssociation
from app.models.player import Player
from app.schemas.player import ImportError as ImportRowError
from app.schemas.player import ImportResult
from app.utils.age import age_category, calculate_age

# Canonical export/import column order.
EXPORT_COLUMNS = [
    "federation_player_code",
    "full_name",
    "gender",
    "date_of_birth",
    "age",
    "age_category",
    "nationality",
    "phone",
    "email",
    "status",
    "club_name",
    "state_name",
]

_GENDER_ALIASES = {
    "m": Gender.male, "male": Gender.male, "man": Gender.male,
    "f": Gender.female, "female": Gender.female, "woman": Gender.female,
    "o": Gender.other, "other": Gender.other, "x": Gender.other,
}


# ------------------------------------------------------------------ helpers
def _lookup_names(db: Session, federation_id: uuid.UUID) -> tuple[dict, dict]:
    clubs = db.execute(
        select(Club.name, Club.id).where(
            Club.federation_id == federation_id, Club.deleted_at.is_(None)
        )
    ).all()
    states = db.execute(
        select(StateAssociation.name, StateAssociation.id).where(
            StateAssociation.federation_id == federation_id,
            StateAssociation.deleted_at.is_(None),
        )
    ).all()
    club_map = {name.strip().lower(): cid for name, cid in clubs}
    state_map = {name.strip().lower(): sid for name, sid in states}
    return club_map, state_map


def _parse_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date '{text}' (use YYYY-MM-DD)")


def _parse_gender(value: object) -> Gender:
    key = str(value or "").strip().lower()
    if key not in _GENDER_ALIASES:
        raise ValueError(f"Unknown gender '{value}' (use M/F/O)")
    return _GENDER_ALIASES[key]


def _parse_status(value: object) -> PlayerStatus:
    key = str(value or "").strip().lower()
    if not key:
        return PlayerStatus.active
    try:
        return PlayerStatus(key)
    except ValueError as exc:
        raise ValueError(f"Unknown status '{value}'") from exc


# ------------------------------------------------------------------ export
def _rows_for_export(db: Session, players: list[Player]) -> list[dict]:
    club_ids = {p.club_id for p in players if p.club_id}
    state_ids = {p.state_id for p in players if p.state_id}
    club_names = dict(
        db.execute(select(Club.id, Club.name).where(Club.id.in_(club_ids))).all()
    ) if club_ids else {}
    state_names = dict(
        db.execute(
            select(StateAssociation.id, StateAssociation.name).where(
                StateAssociation.id.in_(state_ids)
            )
        ).all()
    ) if state_ids else {}
    rows = []
    for p in players:
        rows.append(
            {
                "federation_player_code": p.federation_player_code,
                "full_name": p.full_name,
                "gender": p.gender.value if isinstance(p.gender, Gender) else p.gender,
                "date_of_birth": p.date_of_birth.isoformat() if p.date_of_birth else "",
                "age": calculate_age(p.date_of_birth) or "",
                "age_category": age_category(p.date_of_birth) or "",
                "nationality": p.nationality or "",
                "phone": p.phone or "",
                "email": p.email or "",
                "status": p.status.value if isinstance(p.status, PlayerStatus) else p.status,
                "club_name": club_names.get(p.club_id, ""),
                "state_name": state_names.get(p.state_id, ""),
            }
        )
    return rows


def export_csv(db: Session, players: list[Player]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    writer.writerows(_rows_for_export(db, players))
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx(db: Session, players: list[Player]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Players"
    ws.append(EXPORT_COLUMNS)
    for row in _rows_for_export(db, players):
        ws.append([row[col] for col in EXPORT_COLUMNS])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ------------------------------------------------------------------ import
def _read_records(content: bytes, filename: str) -> list[dict]:
    """Return list of dict rows from a CSV or XLSX upload."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(rows)]
        except StopIteration:
            return []
        records = []
        for values in rows:
            if values is None or all(v is None for v in values):
                continue
            records.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
        return records
    # CSV (default)
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def import_players(
    db: Session, content: bytes, filename: str, *, federation_id: uuid.UUID
) -> ImportResult:
    records = _read_records(content, filename)
    club_map, state_map = _lookup_names(db, federation_id)

    created = updated = skipped = 0
    errors: list[ImportRowError] = []

    for idx, raw in enumerate(records, start=2):  # row 1 = header
        row = {(k or "").strip().lower(): v for k, v in raw.items()}
        code = str(row.get("federation_player_code") or "").strip()
        full_name = str(row.get("full_name") or "").strip()
        if not code or not full_name:
            skipped += 1
            errors.append(
                ImportRowError(row=idx, message="Missing federation_player_code or full_name")
            )
            continue
        try:
            gender = _parse_gender(row.get("gender"))
            dob = _parse_date(row.get("date_of_birth"))
            status = _parse_status(row.get("status"))
        except ValueError as exc:
            skipped += 1
            errors.append(ImportRowError(row=idx, message=str(exc)))
            continue

        club_id = club_map.get(str(row.get("club_name") or "").strip().lower())
        state_id = state_map.get(str(row.get("state_name") or "").strip().lower())

        existing = db.execute(
            select(Player).where(
                Player.federation_id == federation_id,
                Player.federation_player_code == code,
                Player.deleted_at.is_(None),
            )
        ).scalar_one_or_none()

        fields = dict(
            full_name=full_name,
            gender=gender,
            date_of_birth=dob,
            nationality=(str(row.get("nationality")).strip() if row.get("nationality") else None),
            phone=(str(row.get("phone")).strip() if row.get("phone") else None),
            email=(str(row.get("email")).strip() if row.get("email") else None),
            status=status,
            club_id=club_id,
            state_id=state_id,
        )
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
            updated += 1
        else:
            db.add(Player(federation_id=federation_id, federation_player_code=code, **fields))
            created += 1
        # Flush so a later row with the same code in this file resolves as an update
        # instead of colliding on the unique (federation_id, code) constraint.
        db.flush()

    db.commit()
    return ImportResult(created=created, updated=updated, skipped=skipped, errors=errors)


def player_count(db: Session, federation_id: uuid.UUID | None) -> int:
    stmt = select(func.count()).select_from(Player).where(Player.deleted_at.is_(None))
    if federation_id is not None:
        stmt = stmt.where(Player.federation_id == federation_id)
    return db.execute(stmt).scalar_one()
